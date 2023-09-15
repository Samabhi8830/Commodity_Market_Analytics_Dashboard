import xlwings as xw
import openpyxl
import pandas as pd
import dateutil.parser
import eikon as ek
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side
from openpyxl.styles.colors import Color
from datetime import time
import datetime
import pytz
import xlwings as xw
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import Rule
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import ColorScale
from openpyxl.styles import PatternFill, Color

import numpy as np
# Open the Excel workbook
# Set your API key
ek.set_app_key('7c4f90c2de7b42d89076181aee296ca086f79f2c')

wb = xw.Book('trial.xlsx')

# Select the worksheet you want to work with
sheet = wb.sheets['Sheet1']

# Define the cell to monitor for instructions
instruction_cell = sheet.range('G1')
instrument = sheet.range('A1')
start_date = sheet.range('B1')
end_date = sheet.range('C1')
output_interval = sheet.range('D1')
start_time_hour = sheet.range('E1')
end_time_hour = sheet.range('F1')
start_time_min = sheet.range('E2')
end_time_min = sheet.range('F2')
# Define the function to execute the Python code

counter = 0


def generate_heatmaps():
    # Get the user input from the Excel sheet
    com_name = instru[:2]
    Dict = {'1S': 0.25, '1C': 0.25, 'SB': 0.01,
            'CT': 0.01, '1SM': 0.1, '1B': 0.01, '1W': 0.25, '1K': 0.25, 'LR': 1, 'BL': 0.25, 'KC': 0.05, 'LC': 0.01, 'LCC': 1, 'LS': 0.1, 'LG': 0.25, 'CL': 0.01, 'HO': 0.0001}
    ticks = Dict[com_name]

    if Dict[com_name] == '1S':
        com_name1 = instru[:3]
        if com_name1 == '1SM':
            ticks = Dict[com_name1]
    if Dict[com_name] == 'LC':
        com_name1 = instru[:3]
        if com_name1 == 'LCC':
            ticks = Dict[com_name1]

    if __name__ == "__main__":
        data = ek.get_timeseries(instru, start_date=start,
                                 end_date=end, interval="minute")
        Df = pd.DataFrame(data)
        print(Df)
        Df.index = Df.index - pd.to_timedelta('1 minute')
        print(Df)
        Df.index = pd.to_datetime(
            Df.index, format="%d-%m-%Y %H:%M:%S") + datetime.timedelta(hours=5, minutes=30)

        print(Df)
        # start_time = datetime.time(19, 00)
        # end_time = datetime.time(23, 59)
        # df = df.loc[
        #     (df["Date"] >= pd.to_datetime(start_date_value).date()) &
        #     (df["Date"] <= pd.to_datetime(end_date_value).date()) &
        #     (df["Time"] >= start_time) &
        #     (df["Time"] <= end_time)
        # ]

        df = Df.resample(interval).agg({
            "HIGH": "max",
            "LOW": "min",
            "OPEN": "first",
            "CLOSE": "last",
            "VOLUME": "sum"
        })
        print(df)
        df.dropna(inplace=True)
        df["DateTime"] = pd.to_datetime(
            df.index, format="%d-%m-%Y %H:%M:%S")
        df["Time"] = df["DateTime"].dt.time
        df["Dates"] = df["DateTime"].dt.date
        print(df)
        # Filter the data based on the user-defined start and end dates and time range
        start_date_value = start
        end_date_value = end

        start_time = datetime.time(int(sTh), int(sTm))
        end_time = datetime.time(int(eTh), int(eTm))
        df = df.loc[
            (df["Dates"] >= pd.to_datetime(start_date_value).date()) &
            (df["Dates"] <= pd.to_datetime(end_date_value).date()) &
            (df["Time"] >= start_time) &
            (df["Time"] <= end_time)
        ]
        print(df)
        # df.head()
        df['Dates'] = pd.to_datetime(df['Dates'], errors='coerce')

        # Exclude weekends (Saturday and Sunday)
        df = df.loc[df["Dates"].dt.weekday < 5]

        # Pivot the DataFrame to get the desired structure for the heat map
        df_heatmap_volume = df.pivot(
            index="Dates", columns="Time", values="VOLUME")
        df_heatmap_open = df.pivot(
            index="Dates", columns="Time", values="OPEN")
        df_heatmap_close = df.pivot(
            index="Dates", columns="Time", values="CLOSE")
        df_heatmap_high = df.pivot(
            index="Dates", columns="Time", values="HIGH")
        df_heatmap_low = df.pivot(index="Dates", columns="Time", values="LOW")
        df_heatmap_open_close_diff = df.pivot(index="Dates", columns="Time", values="CLOSE") - df.pivot(
            index="Dates", columns="Time", values="OPEN")
        df_heatmap_high_low_diff = df.pivot(
            index="Dates", columns="Time", values="HIGH") - df.pivot(index="Dates", columns="Time", values="LOW")

        # Create the heat maps using Matplotlib and Seaborn
        heatmaps = {
            "Volume": df_heatmap_volume,
            "Range": df_heatmap_high_low_diff,
            "Change": df_heatmap_open_close_diff,
            "Open": df_heatmap_open,
            "Close": df_heatmap_close,
            "High": df_heatmap_high,
            "Low": df_heatmap_low
        }

        # Set the initial row and column indices
        row_index = 3
        col_index = 2
        counter = 0
        # Iterate over the heatmaps and insert them on the same sheet
        for sheet_name, df_heatmap in heatmaps.items():
            # Write the name of the heatmap above it
            sheet.range(row_index, col_index -
                        1).value = sheet_name
            df_heatmap = df_heatmap.fillna(0)

            for row_idx, (date_value, row_data) in enumerate(df_heatmap.iterrows(), start=row_index + 1):
                for col_idx, value in enumerate(row_data, start=col_index):
                    if pd.notna(value):
                        if Dict[com_name] == 0.0001:
                            if counter == 0:
                                sheet.range(
                                    row_idx, col_idx).value = format(value, ".2f")
                            elif counter == 1 or counter == 2:
                                sheet.range(
                                    row_idx, col_idx).value = format(value/ticks, ".4f")
                            else:
                                sheet.range(
                                    row_idx, col_idx).value = format(value, ".4f")

                        else:
                            if counter == 1 or counter == 2:
                                sheet.range(
                                    row_idx, col_idx).value = format(value/ticks, ".2f")
                            else:
                                sheet.range(
                                    row_idx, col_idx).value = format(value, ".2f")

                    else:
                        sheet.range(
                            row_idx, col_idx).value = 0

            # Write the date values in the first column
            date_values = df_heatmap.index.date.tolist()

            for row_idx, date_value in enumerate(date_values, start=row_index + 1):
                sheet.range(row_idx, 1).value = date_value

            # Write the time values in the first row
            time_values = [time.strftime(col, "%H:%M:%S")
                           for col in df_heatmap.columns]

            for col_idx, time_value in enumerate(time_values, start=col_index):
                sheet.range(row_index, col_idx).value = time_value

            valid_values = df_heatmap.values.flatten()
            valid_values = valid_values[~pd.isna(valid_values)]

            min_value = np.nanmin(valid_values)
            max_value = np.nanmax(valid_values)

            # Calculate and add the average row
            average_row_index = row_index + df_heatmap.shape[0] + 1

            if counter == 0:
                sheet.range(average_row_index, col_index -
                            1).value = "Sum of Abs"
                for col_idx in range(col_index, col_index + df_heatmap.shape[1]):
                    average_value = ((df_heatmap.iloc[:,
                                                      col_idx - col_index]).mean())
                    cell = sheet.range(average_row_index, col_idx)
                    cell.value = format(average_value, ".2f")
            elif counter == 1 or counter == 2:
                sheet.range(average_row_index, col_index -
                            1).value = "Sum of Abs"
                for col_idx in range(col_index, col_index + df_heatmap.shape[1]):
                    average_value = ((df_heatmap.iloc[:,
                                                      col_idx - col_index] / ticks).abs().sum()/df_heatmap.shape[0])
                    cell = sheet.range(average_row_index, col_idx)
                    cell.value = format(average_value, ".2f")
            else:
                sheet.range(average_row_index, col_index - 1).value = "Average"
                for col_idx in range(col_index, col_index + df_heatmap.shape[1]):
                    average_value = df_heatmap.iloc[:,
                                                    col_idx - col_index].mean()
                    cell = sheet.range(average_row_index, col_idx)
                    cell.value = format(average_value, ".2f")

            for row_number in range(row_index + 1, row_index + df_heatmap.shape[0]+2):

                for cell in sheet.range((row_number, col_index), (row_number, col_index + df_heatmap.shape[1]-1)):

                    value = cell.value
                    # Convert the cell value to a numpy array
                    arr = np.array(value)

                # Calculate the color based on the value (arr) or any other logic
                    min_value = np.min(
                        df_heatmap.iloc[1:, 1:].astype(int))
                    max_value = np.max(
                        df_heatmap.iloc[1:, 1:].astype(int))
                    normalized_value = (arr - min_value) / \
                        (max_value - min_value)

                    if normalized_value[0] <= 0 or value <= 0:

                        red_intensity = np.uint8(normalized_value * 255)
                        green_intensity = np.uint8(
                            (1 - normalized_value) * 255)

    # Apply the color to the cell using red and green intensities
                        cell.color = (255, int(
                            green_intensity[0]), 0)
                    else:
                        red_intensity = np.uint8(normalized_value * 255)
                        green_intensity = np.uint8(
                            (1 - normalized_value) * 255)

    # Apply the color to the cell using red and green intensities
                        cell.color = (int(
                            red_intensity[0]), 255, 0)

                    arr.fill(0)

            max_rows_value = -10000
            min_rows_value = 10000
            for row_number in range(row_index + 1, row_index + df_heatmap.shape[0]+2):

                max_value = 0
                for cell in sheet.range((row_number, col_index), (row_number, col_index + df_heatmap.shape[1]-1)):

                    cell_value = cell.value
                    cell_value1 = cell.value

                    if cell_value != None and cell_value > max_rows_value:
                        max_rows_value = cell_value
                        max_cell_address = cell.address

                    if cell_value1 != None and cell_value1 < min_rows_value:
                        min_rows_value = cell_value1
                        min_cell_address = cell.address

                cell_max = sheet.range(max_cell_address)

                cell_max.api.Borders.LineStyle = xw.constants.LineStyle.xlContinuous
                cell_max.api.Borders.Weight = 4
                cell_max.api.Borders.Color = xw.utils.rgb_to_int(
                    (0, 0, 0))  # Black color
                cello = cell_max
                cello.color = (0, 255, 0)
                cell_max = 0
                max_rows_value = -10000
                cell_min = sheet.range(min_cell_address)
                cello1 = cell_min
                cell_min.api.Borders.LineStyle = xw.constants.LineStyle.xlContinuous
                cell_min.api.Borders.Weight = 4
                cell_min.api.Borders.Color = xw.utils.rgb_to_int(
                    (0, 0, 0))  # Black color
                cello1.color = (255, 0, 0)
                cell_min = 0
                min_rows_value = 10000

            # Update the row index for the next heatmap
            row_index += df_heatmap.shape[0] + 4
            counter += 1

    else:
        sheet["H1"] = "Invalid input format. Please enter correct values."


# Continuously monitor the cell and execute the code when the instruction is entered
while True:
    # Read the instruction from the cell
    instruction = instruction_cell.value
    instru = instrument.value
    start = start_date.value
    end = end_date.value
    interval = output_interval.value
    sTh = start_time_hour.value
    sTm = start_time_min.value
    eTh = end_time_hour.value
    eTm = end_time_min.value
    # Check if the instruction is "1" to run the code
    if instruction == 1:
        # Clear the instruction cell
        used_range = sheet.used_range
        used_range.clear()
        try:
            # Check if the instrument is valid
            data = ek.get_timeseries(
                instru, start_date=start, end_date=end, interval="minute")
        except ek.EikonError as e:
            # Display error message if instrument is invalid
            sheet.range(
                'H1').value = 'Invalid instrument. Please enter a valid instrument code.'
            sheet.range('A2').value = instru
            sheet.range('B2').value = start
            sheet.range('C2').value = end
            sheet.range('D2').value = interval
            sheet.range('G2').value = instruction
            sheet.range('E1').value = sTh
            sheet.range('F1').value = eTh
            sheet.range('E2').value = sTm
            sheet.range('F2').value = eTm
            continue
        except dateutil.parser.ParserError as e:
            # Display error message if date format is invalid
            sheet.range(
                'H2').value = 'Invalid date format. Please enter the date in YYYY-MM-DD format.'
            sheet.range('A2').value = instru
            sheet.range('B2').value = start
            sheet.range('C2').value = end
            sheet.range('D2').value = interval
            sheet.range('G2').value = instruction
            sheet.range('E1').value = sTh
            sheet.range('F1').value = eTh
            sheet.range('E2').value = sTm
            sheet.range('F2').value = eTm
            continue

        try:
            # Start time after 19:00
            start_time = datetime.time(int(sTh), int(sTm))
            end_time = datetime.time(int(eTh), int(eTm))
            # Rest of your code using start_time and end_time variables
        except ValueError as e:
            sheet.range(
                'I1').value = "Error: Invalid input for time"
            sheet.range('A2').value = instru
            sheet.range('B2').value = start
            sheet.range('C2').value = end
            sheet.range('D2').value = interval
            sheet.range('G2').value = instruction
            sheet.range('E1').value = sTh
            sheet.range('F1').value = eTh
            sheet.range('E2').value = sTm
            sheet.range('F2').value = eTm
            continue
            # Handle the specific exception, such as displaying an error message or taking alternative actions
        except Exception as e:
            sheet.range(
                'I1').value = "Error: Invalid input for time"
            continue
            # Handle any other exceptions that might occur

        sheet.range('A2').value = instru
        sheet.range('B2').value = start
        sheet.range('C2').value = end
        sheet.range('D2').value = interval
        sheet.range('G2').value = instruction
        sheet.range('E1').value = sTh
        sheet.range('F1').value = eTh
        sheet.range('E2').value = sTm
        sheet.range('F2').value = eTm
        # print(sT)
        # print(start_date)
        instruction_cell.value = ''
        instrument.value = ''
        start_date.value = ''
        end_date.value = ''
        output_interval.value = ''
        # start_time_hour.value = ''
        # end_time_hour.value = ''
        # start_time_min.value = ''
        # end_time_min.value = ''

        generate_heatmaps()
# Close the workbook
wb.close()
